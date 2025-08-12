#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
Excel文件搜索引擎模块
"""

import os
import re
import time
from datetime import datetime
from PyQt6.QtCore import QObject, pyqtSignal
import pandas as pd
import openpyxl
import xlrd
from .utils.logger import get_logger

logger = get_logger(__name__)

class SearchEngine(QObject):
    """高性能Excel文件搜索引擎"""
    
    # 信号定义
    file_found = pyqtSignal(dict)  # 当找到包含关键字的文件时发出
    search_progress = pyqtSignal(int, int)  # 搜索进度信号，参数：当前进度，总数
    search_finished = pyqtSignal(int, int)  # 搜索完成信号，参数：总文件数，找到的文件数
    search_error = pyqtSignal(str)  # 搜索错误信号
    
    def __init__(self):
        super().__init__()
        self.directory = ""  # 搜索目录
        self.keyword = ""  # 搜索关键字
        self.case_sensitive = False  # 是否区分大小写
        self.whole_word = False  # 是否完整单词匹配
        self.include_subdirs = True  # 是否包含子目录
        self.file_type = "All Excel Files (.xlsx, .xls)"  # 文件类型过滤
        self.stop_flag = False  # 停止搜索标志
        self.buffer_size = 10 * 1024 * 1024  # 缓冲区大小（10MB）- 增加以支持大量Excel文件
        self.complete_search = True  # 是否完整搜索（True=搜索所有行，False=只搜索前1000行）
        
    def set_search_params(self, directory, keyword, case_sensitive=False, 
                         whole_word=False, include_subdirs=True, file_type="All Excel Files (.xlsx, .xls)",
                         complete_search=True):
        """设置搜索参数"""
        self.directory = directory
        self.keyword = keyword
        self.case_sensitive = case_sensitive
        self.whole_word = whole_word
        self.include_subdirs = include_subdirs
        self.file_type = file_type
        self.complete_search = complete_search
        
    def start_search(self):
        """开始搜索过程"""
        try:
            logger.info(f"开始搜索关键字 '{self.keyword}' 在目录 '{self.directory}' 中")
            
            # 构建搜索模式
            pattern = self._build_search_pattern()
            
            # 获取文件列表
            files = self._get_excel_files()
            total_files = len(files)
            found_files = 0
            
            logger.info(f"找到 {total_files} 个Excel文件需要搜索")
            
            # 内存优化：对于大量文件，分批处理
            if total_files > 1000:
                logger.info(f"检测到大量文件({total_files}个)，启用分批处理模式")
                batch_size = 100  # 每批处理100个文件
                for i in range(0, total_files, batch_size):
                    if self.stop_flag:
                        break
                    batch_files = files[i:i + batch_size]
                    found_files += self._process_file_batch(batch_files, pattern, i, total_files)
            else:
                # 小量文件直接处理
                found_files = self._process_file_batch(files, pattern, 0, total_files)
            
            # 发出最终进度和结果
            self.search_progress.emit(total_files, total_files)
            self.search_finished.emit(total_files, found_files)
            
        except Exception as e:
            logger.error(f"搜索错误: {str(e)}")
            self.search_error.emit(str(e))
            
    def _process_file_batch(self, files, pattern, start_index, total_files):
        """分批处理文件"""
        found_files = 0
        
        for i, file_path in enumerate(files):
            if self.stop_flag:
                logger.info("用户停止了搜索")
                break
                
            try:
                # 更新进度
                current_progress = start_index + i
                self.search_progress.emit(current_progress, total_files)
                
                # 搜索文件
                file_info = self._search_file(file_path, pattern)
                if file_info:
                    self.file_found.emit(file_info)
                    found_files += 1
                    
            except Exception as e:
                logger.error(f"搜索文件 {file_path} 时出错: {str(e)}")
                continue
                
        return found_files
            
    def stop_search(self):
        """停止当前搜索"""
        self.stop_flag = True
        
    def _build_search_pattern(self):
        """构建关键字搜索的正则表达式模式"""
        keyword = re.escape(self.keyword)
        
        if self.whole_word:
            pattern = r'\b' + keyword + r'\b'
        else:
            pattern = keyword
            
        if not self.case_sensitive:
            pattern = re.compile(pattern, re.IGNORECASE)
        else:
            pattern = re.compile(pattern)
            
        return pattern
        
    def _get_excel_files(self):
        """获取要搜索的Excel文件列表"""
        files = []
        
        if self.include_subdirs:
            for root, dirs, filenames in os.walk(self.directory):
                for filename in filenames:
                    if self._is_excel_file(filename):
                        files.append(os.path.join(root, filename))
        else:
            for filename in os.listdir(self.directory):
                if self._is_excel_file(filename):
                    files.append(os.path.join(self.directory, filename))
                    
        return files
        
    def _is_excel_file(self, filename):
        """根据当前文件类型过滤器检查文件是否为Excel文件"""
        filename_lower = filename.lower()
        
        if self.file_type == "All Excel Files (.xlsx, .xls)":
            return filename_lower.endswith(('.xlsx', '.xls'))
        elif self.file_type == "Excel 2007+ (.xlsx)":
            return filename_lower.endswith('.xlsx')
        elif self.file_type == "Excel 97-2003 (.xls)":
            return filename_lower.endswith('.xls')
        elif self.file_type == "All Files":
            return filename_lower.endswith(('.xlsx', '.xls', '.csv', '.txt'))
        else:
            return filename_lower.endswith(('.xlsx', '.xls'))
            
    def _search_file(self, file_path, pattern):
        """在单个Excel文件中搜索关键字"""
        try:
            file_info = {
                'name': os.path.basename(file_path),
                'path': file_path,
                'size': self._format_file_size(os.path.getsize(file_path)),
                'modified': datetime.fromtimestamp(os.path.getmtime(file_path)).strftime('%Y-%m-%d %H:%M:%S'),
                'matches': 0,
                'preview': ''
            }
            
            matches = 0
            preview_lines = []
            
            # 首先尝试用pandas读取（对大文件更快）
            try:
                if file_path.lower().endswith('.xlsx'):
                    # 用pandas读取Excel
                    excel_file = pd.ExcelFile(file_path)
                    for sheet_name in excel_file.sheet_names:
                        if self.stop_flag:
                            break
                            
                        try:
                            # 根据配置决定是否完整搜索
                            if self.complete_search:
                                df = pd.read_excel(file_path, sheet_name=sheet_name)  # 完整搜索
                            else:
                                df = pd.read_excel(file_path, sheet_name=sheet_name, nrows=1000)  # 快速搜索
                            sheet_matches = self._search_dataframe(df, pattern, preview_lines)
                            matches += sheet_matches
                            # 及时释放内存
                            del df
                        except Exception as e:
                            logger.warning(f"读取工作表 {sheet_name} 时出错: {str(e)}")
                            continue
                            
                elif file_path.lower().endswith('.xls'):
                    # 读取旧版Excel格式
                    matches = self._search_xls_file(file_path, pattern, preview_lines)
                    
            except Exception as e:
                logger.warning(f"pandas读取 {file_path} 失败，尝试使用openpyxl: {str(e)}")
                # 回退到openpyxl
                matches = self._search_with_openpyxl(file_path, pattern, preview_lines)
                
            if matches > 0:
                file_info['matches'] = matches
                file_info['preview'] = '\n'.join(preview_lines[:10])  # 限制预览行数
                return file_info
                
        except Exception as e:
            logger.error(f"搜索文件 {file_path} 时出错: {str(e)}")
            
        return None
        
    def _search_dataframe(self, df, pattern, preview_lines):
        """在pandas DataFrame中搜索关键字"""
        matches = 0
        
        # 将DataFrame转换为字符串并搜索
        df_str = df.to_string()
        found_matches = pattern.findall(df_str)
        matches += len(found_matches)
        
        # 添加预览行
        if matches > 0:
            lines = df_str.split('\n')
            for line in lines:
                if pattern.search(line):
                    preview_lines.append(line.strip())
                    if len(preview_lines) >= 5:  # 限制预览行数
                        break
                        
        return matches
        
    def _search_xls_file(self, file_path, pattern, preview_lines):
        """搜索旧版Excel (.xls) 文件"""
        matches = 0
        
        try:
            # 首先尝试用xlrd读取
            workbook = xlrd.open_workbook(file_path)
            for sheet in workbook.sheets():
                if self.stop_flag:
                    break
                    
                # 根据配置决定搜索范围
                max_rows = sheet.nrows if self.complete_search else min(sheet.nrows, 1000)
                for row_idx in range(max_rows):
                    row_data = sheet.row_values(row_idx)
                    row_str = ' '.join(str(cell) for cell in row_data)
                    
                    if pattern.search(row_str):
                        matches += len(pattern.findall(row_str))
                        preview_lines.append(row_str.strip())
                        if len(preview_lines) >= 5:
                            break
                            
        except Exception as e:
            logger.warning(f"xlrd读取XLS文件 {file_path} 失败，尝试使用openpyxl: {str(e)}")
            # 如果xlrd失败，尝试用openpyxl读取
            try:
                matches = self._search_with_openpyxl(file_path, pattern, preview_lines)
            except Exception as e2:
                logger.error(f"openpyxl读取XLS文件 {file_path} 也失败: {str(e2)}")
            
        return matches
        
    def _search_with_openpyxl(self, file_path, pattern, preview_lines):
        """使用openpyxl作为回退方案进行搜索"""
        matches = 0
        
        try:
            workbook = openpyxl.load_workbook(file_path, read_only=True, data_only=True)
            
            for sheet_name in workbook.sheetnames:
                if self.stop_flag:
                    break
                    
                sheet = workbook[sheet_name]
                # 根据配置决定搜索范围
                max_row = None if self.complete_search else 1000
                for row in sheet.iter_rows(max_row=max_row):
                    row_str = ' '.join(str(cell.value or '') for cell in row)
                    
                    if pattern.search(row_str):
                        matches += len(pattern.findall(row_str))
                        preview_lines.append(row_str.strip())
                        if len(preview_lines) >= 5:
                            break
                            
            workbook.close()
            
        except Exception as e:
            logger.error(f"使用openpyxl读取 {file_path} 时出错: {str(e)}")
            
        return matches
        
    def _format_file_size(self, size_bytes):
        """格式化文件大小为人类可读格式"""
        if size_bytes == 0:
            return "0 B"
            
        size_names = ["B", "KB", "MB", "GB"]
        i = 0
        while size_bytes >= 1024 and i < len(size_names) - 1:
            size_bytes /= 1024.0
            i += 1
            
        return f"{size_bytes:.1f} {size_names[i]}"
